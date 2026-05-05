#!/usr/bin/env python

import logging
import tempfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from click.testing import CliRunner

from crosstab import __version__
from crosstab.cli import app
from crosstab.crosstab import Crosstab, _quote_ident

runner = CliRunner()

logger = logging.getLogger(__name__)

# Sample data for tests
CSV_CONTENT = """header1,header2,header3,value,unit
A,1,2018,10,%
A,1,2019,20,%
B,2,2018,30,%
B,2,2019,40,%
"""

CSV_DUPLICATE_CONTENT = """header1,header2,header3,value,unit
A,1,2018,10,%
A,1,2018,11,%
B,2,2019,40,%
"""


@pytest.fixture
def temp_csv_file():
    """Fixture to create a temporary CSV file"""
    with tempfile.NamedTemporaryFile(mode="w+", suffix=".csv", delete=False) as f:
        f.write(CSV_CONTENT)
        f.seek(0)
        yield Path(f.name)
    Path(f.name).unlink()  # Clean up


@pytest.fixture
def temp_csv_with_duplicates():
    """Fixture for a CSV file containing duplicate row/col combinations."""
    with tempfile.NamedTemporaryFile(mode="w+", suffix=".csv", delete=False) as f:
        f.write(CSV_DUPLICATE_CONTENT)
        f.seek(0)
        yield Path(f.name)
    Path(f.name).unlink()


@pytest.fixture
def temp_xlsx_file():
    """Fixture to create a temporary XLSX file"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        yield Path(f.name)
    Path(f.name).unlink()  # Clean up


def test_validate_args(temp_csv_file, temp_xlsx_file):
    """Test the validation of arguments."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2",),
        value_cols=("value", "unit"),
    )
    assert crosstab.incsv == temp_csv_file
    assert crosstab.outxlsx == temp_xlsx_file


def test_default_output_path(temp_csv_file):
    """When outxlsx is not provided, default to <stem>_crosstab.xlsx alongside the input."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        row_headers=("header1",),
        col_headers=("header2",),
        value_cols=("value",),
    )
    assert crosstab.outxlsx == temp_csv_file.with_name(temp_csv_file.stem + "_crosstab.xlsx")


def test_repr_contains_paths(temp_csv_file, temp_xlsx_file):
    """__repr__ should contain the configured paths and headers."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2",),
        value_cols=("value",),
    )
    rep = repr(crosstab)
    assert "Crosstab(" in rep
    # Compare via the path's repr — on Windows ``WindowsPath.__repr__``
    # double-escapes backslashes so ``str(path)`` is not a substring of
    # the surrounding repr.
    assert repr(temp_csv_file) in rep
    assert repr(temp_xlsx_file) in rep
    assert "header1" in rep
    assert "header2" in rep
    assert "value" in rep


def test_invalid_args_missing_file():
    """Test with missing CSV file"""
    with pytest.raises(ValueError, match="Input file .* does not exist."):
        Crosstab(
            incsv=Path("missing.csv"),
            outxlsx=Path("output.xlsx"),
            row_headers=("header1",),
            col_headers=("header2",),
            value_cols=("value", "unit"),
        )


def test_invalid_args_not_a_file(tmp_path):
    """Test with a directory passed as incsv."""
    directory = tmp_path / "not_a_file.csv"
    directory.mkdir()
    with pytest.raises(ValueError, match="is not a file"):
        Crosstab(
            incsv=directory,
            outxlsx=Path("output.xlsx"),
            row_headers=("header1",),
            col_headers=("header2",),
            value_cols=("value",),
        )


def test_invalid_args_empty_file():
    """Test with an empty CSV file."""
    # Create an empty temporary CSV file
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
        temp_csv = Path(f.name)
    with pytest.raises(ValueError, match="Input file .* is empty."):
        Crosstab(
            incsv=temp_csv,
            outxlsx=Path("output.xlsx"),
            row_headers=("header1",),
            col_headers=("header2",),
            value_cols=("value", "unit"),
        )
    temp_csv.unlink()


def test_invalid_args_wrong_csv_extension(tmp_path):
    """Input must have a .csv extension."""
    bad = tmp_path / "input.txt"
    bad.write_text("a,b\n1,2\n")
    with pytest.raises(ValueError, match="is not a CSV file"):
        Crosstab(
            incsv=bad,
            outxlsx=Path("output.xlsx"),
            row_headers=("a",),
            col_headers=("b",),
            value_cols=("a",),
        )


def test_invalid_args_wrong_xlsx_extension(temp_csv_file, tmp_path):
    """Output must have a .xlsx extension."""
    bad = tmp_path / "output.csv"
    with pytest.raises(ValueError, match="must have an XLSX extension"):
        Crosstab(
            incsv=temp_csv_file,
            outxlsx=bad,
            row_headers=("header1",),
            col_headers=("header2",),
            value_cols=("value",),
        )


def test_invalid_args_empty_row_headers(temp_csv_file, temp_xlsx_file):
    with pytest.raises(ValueError, match="No row headers specified."):
        Crosstab(
            incsv=temp_csv_file,
            outxlsx=temp_xlsx_file,
            row_headers=(),
            col_headers=("header2",),
            value_cols=("value",),
        )


def test_invalid_args_empty_col_headers(temp_csv_file, temp_xlsx_file):
    with pytest.raises(ValueError, match="No column headers specified."):
        Crosstab(
            incsv=temp_csv_file,
            outxlsx=temp_xlsx_file,
            row_headers=("header1",),
            col_headers=(),
            value_cols=("value",),
        )


def test_invalid_args_empty_value_cols(temp_csv_file, temp_xlsx_file):
    with pytest.raises(ValueError, match="No value columns specified."):
        Crosstab(
            incsv=temp_csv_file,
            outxlsx=temp_xlsx_file,
            row_headers=("header1",),
            col_headers=("header2",),
            value_cols=(),
        )


def test_invalid_args_unknown_header(temp_csv_file, temp_xlsx_file):
    """Requesting a column that does not exist in the CSV should raise."""
    with pytest.raises(ValueError, match="Headers not found in CSV file"):
        Crosstab(
            incsv=temp_csv_file,
            outxlsx=temp_xlsx_file,
            row_headers=("not_a_real_column",),
            col_headers=("header2",),
            value_cols=("value",),
        )


def test_csv_columns_are_detected(temp_csv_file):
    """The Crosstab constructor reads the CSV header row via DuckDB."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        outxlsx=Path("output.xlsx"),
        row_headers=("header1",),
        col_headers=("header2",),
        value_cols=("value", "unit"),
    )
    assert crosstab.csv_columns == ("header1", "header2", "header3", "value", "unit")
    crosstab.close()


def test_keep_duckdb_creates_database_file(tmp_path):
    """`keep_duckdb=True` produces a queryable DuckDB file alongside the input."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    out_path = tmp_path / "out.xlsx"
    duckdb_path = csv_path.with_suffix(".duckdb")

    Crosstab(
        incsv=csv_path,
        outxlsx=out_path,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value",),
        keep_duckdb=True,
    ).crosstab()

    assert out_path.exists()
    assert duckdb_path.exists()
    assert duckdb_path.stat().st_size > 0


def test_keep_duckdb_overwrites_existing_database_file(tmp_path):
    """A pre-existing .duckdb at the target path is replaced, not appended."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    duckdb_path = csv_path.with_suffix(".duckdb")
    duckdb_path.write_bytes(b"stale-bytes")

    Crosstab(
        incsv=csv_path,
        outxlsx=tmp_path / "out.xlsx",
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value",),
        keep_duckdb=True,
    ).crosstab()

    assert duckdb_path.read_bytes() != b"stale-bytes"


def test_fill_substitutes_blanks(temp_csv_file, temp_xlsx_file):
    """When `fill` is set, missing (row, col) pairs render with the fill value."""
    Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value",),
        fill="N/A",
    ).crosstab()
    wb = openpyxl.load_workbook(temp_xlsx_file)
    ws = wb["Crosstab"]
    # In the fixture, (A, 2, 2018) and (A, 2, 2019) have no rows, so those
    # cells should now be filled.
    assert ws["D4"].value == "N/A"
    assert ws["E4"].value == "N/A"
    # Cells with real data are unaffected.
    assert ws["B4"].value == "10"
    wb.close()


def test_duplicate_row_col_combination_raises(temp_csv_with_duplicates, temp_xlsx_file):
    """When the same row/col combination appears twice, crosstab() should raise."""
    crosstab = Crosstab(
        incsv=temp_csv_with_duplicates,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value",),
    )
    with pytest.raises(ValueError, match="Multiple values found"):
        crosstab.crosstab()


def test_crosstab_creation(temp_csv_file, temp_xlsx_file):
    """Test the creation of a crosstab file."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value", "unit"),
        keep_src=True,
    )
    crosstab.crosstab()
    assert temp_xlsx_file.exists()
    # Crosstab + Source Data — no README sheet anymore.
    wb = openpyxl.load_workbook(temp_xlsx_file)
    assert wb.sheetnames == ["Crosstab", "Source Data"]


def test_crosstab_omits_source_sheet_when_keep_src_false(temp_csv_file, temp_xlsx_file):
    """With keep_src=False the workbook contains only the Crosstab sheet."""
    Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value",),
        keep_src=False,
    ).crosstab()
    wb = openpyxl.load_workbook(temp_xlsx_file)
    assert wb.sheetnames == ["Crosstab"]


def test_crosstab_rows_single_value_column(temp_csv_file, temp_xlsx_file):
    """Test that values are correctly placed in the crosstab with one value column."""
    Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value",),
        keep_src=True,
    ).crosstab()
    wb = openpyxl.load_workbook(temp_xlsx_file)
    ws = wb["Crosstab"]
    # Check the row headers
    assert ws["A1"].value == "header2"
    assert ws["A2"].value == "header3"
    assert ws["A3"].value == "header1"
    assert ws["A4"].value == "A"
    assert ws["A5"].value == "B"
    # Check the column headers
    assert ws["B1"].value == "1"
    assert ws["C1"].value == "1"
    assert ws["D1"].value == "2"
    assert ws["E1"].value == "2"
    assert ws["B2"].value == "2018"
    assert ws["C2"].value == "2019"
    assert ws["D2"].value == "2018"
    assert ws["E2"].value == "2019"
    assert ws["B3"].value == "value"
    assert ws["C3"].value == "value"
    assert ws["D3"].value == "value"
    assert ws["E3"].value == "value"
    # Check the values
    assert ws["B4"].value == "10"
    assert ws["C4"].value == "20"
    assert ws["D4"].value is None
    assert ws["E4"].value is None
    assert ws["B5"].value is None
    assert ws["C5"].value is None
    assert ws["D5"].value == "30"
    assert ws["E5"].value == "40"
    wb.close()


def test_crosstab_rows_multi_value_column(temp_csv_file, temp_xlsx_file):
    """Test that values are correctly placed in the crosstab with multiple value columns."""
    Crosstab(
        incsv=temp_csv_file,
        outxlsx=Path(temp_xlsx_file),
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value", "unit"),
        keep_src=True,
    ).crosstab()
    wb = openpyxl.load_workbook(temp_xlsx_file)
    ws = wb["Crosstab"]
    # Check the row headers
    assert ws["A1"].value == "header2"
    assert ws["A2"].value == "header3"
    assert ws["A3"].value == "header1"
    assert ws["A4"].value == "A"
    assert ws["A5"].value == "B"
    # Check the column headers
    assert ws["B1"].value == "1"
    assert ws["C1"].value is None
    assert ws["D1"].value == "1"
    assert ws["E1"].value is None
    assert ws["F1"].value == "2"
    assert ws["G1"].value is None
    assert ws["H1"].value == "2"
    assert ws["I1"].value is None
    assert ws["B2"].value == "2018"
    assert ws["C2"].value is None
    assert ws["D2"].value == "2019"
    assert ws["E2"].value is None
    assert ws["F2"].value == "2018"
    assert ws["G2"].value is None
    assert ws["H2"].value == "2019"
    assert ws["I2"].value is None
    assert ws["B3"].value == "value"
    assert ws["C3"].value == "unit"
    assert ws["D3"].value == "value"
    assert ws["E3"].value == "unit"
    assert ws["F3"].value == "value"
    assert ws["G3"].value == "unit"
    assert ws["H3"].value == "value"
    assert ws["I3"].value == "unit"
    # Check the values
    assert ws["B4"].value == "10"
    assert ws["C4"].value == "%"
    assert ws["D4"].value == "20"
    assert ws["E4"].value == "%"
    assert ws["F4"].value is None
    assert ws["G4"].value is None
    assert ws["H4"].value is None
    assert ws["I4"].value is None
    assert ws["B5"].value is None
    assert ws["C5"].value is None
    assert ws["D5"].value is None
    assert ws["E5"].value is None
    assert ws["F5"].value == "30"
    assert ws["G5"].value == "%"
    assert ws["H5"].value == "40"
    assert ws["I5"].value == "%"
    wb.close()


# ── CLI tests ─────────────────────────────────────────────────────────────────


def test_cli_no_args_shows_help():
    """Invoking with no arguments should print the help screen and exit 2."""
    result = runner.invoke(app, [])
    assert result.exit_code == 2
    assert "Usage" in result.output


def test_cli_version_flag():
    """`--version` should print the version and exit 0."""
    result = runner.invoke(app, ["--version"])
    assert result.exit_code == 0
    assert __version__ in result.output


def test_cli_runs_end_to_end(tmp_path):
    """Invoke the CLI with valid args and confirm the output XLSX is produced."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    out_path = tmp_path / "out.xlsx"
    result = runner.invoke(
        app,
        [
            "-q",
            "-s",
            "-f",
            str(csv_path),
            "-o",
            str(out_path),
            "-r",
            "header1",
            "-c",
            "header2",
            "header3",
            "-v",
            "value",
        ],
    )
    assert result.exit_code == 0, result.output
    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert "Crosstab" in wb.sheetnames
    assert "Source Data" in wb.sheetnames
    wb.close()


def test_cli_multi_value_flags_are_space_separated(tmp_path):
    """`-r foo bar` consumes both tokens into row_headers (nargs='+' style)."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    out_path = tmp_path / "out.xlsx"
    result = runner.invoke(
        app,
        [
            "-q",
            "-f",
            str(csv_path),
            "-o",
            str(out_path),
            "-r",
            "header1",
            "-c",
            "header2",
            "header3",
            "-v",
            "value",
            "unit",
        ],
    )
    assert result.exit_code == 0, result.output
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Crosstab"]
    # Two value columns means each column-header block gets two sub-columns,
    # so the value-label row alternates "value" / "unit".
    assert ws["B3"].value == "value"
    assert ws["C3"].value == "unit"
    wb.close()


def test_cli_debug_log_to_file(tmp_path):
    """`--debug` plus `--log` should write debug output to the log file."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    out_path = tmp_path / "out.xlsx"
    log_path = tmp_path / "run.log"
    result = runner.invoke(
        app,
        [
            "-d",
            "-l",
            str(log_path),
            "-f",
            str(csv_path),
            "-o",
            str(out_path),
            "-r",
            "header1",
            "-c",
            "header2",
            "header3",
            "-v",
            "value",
        ],
    )
    assert result.exit_code == 0, result.output
    assert out_path.exists()
    assert log_path.exists()
    assert log_path.read_text()  # non-empty


def test_cli_value_error_exits_nonzero(tmp_path):
    """A ValueError from the engine should produce exit code 1."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_DUPLICATE_CONTENT)
    out_path = tmp_path / "out.xlsx"
    result = runner.invoke(
        app,
        [
            "-q",
            "-f",
            str(csv_path),
            "-o",
            str(out_path),
            "-r",
            "header1",
            "-c",
            "header2",
            "header3",
            "-v",
            "value",
        ],
    )
    assert result.exit_code == 1


def test_cli_unexpected_error_exits_nonzero(tmp_path):
    """An unexpected exception from the engine should produce exit code 1."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    out_path = tmp_path / "out.xlsx"
    with patch("crosstab.cli.Crosstab.crosstab", side_effect=RuntimeError("boom")):
        result = runner.invoke(
            app,
            [
                "-q",
                "-f",
                str(csv_path),
                "-o",
                str(out_path),
                "-r",
                "header1",
                "-c",
                "header2",
                "header3",
                "-v",
                "value",
            ],
        )
    assert result.exit_code == 1


def test_cli_missing_required_input_fails():
    """Omitting --input must fail with a non-zero exit (typer/click validation)."""
    result = runner.invoke(app, ["-r", "a", "-c", "b", "-v", "v"])
    assert result.exit_code != 0


def test_cli_keep_duckdb_creates_database_file(tmp_path):
    """`--keep-duckdb` writes a queryable DuckDB file alongside the input."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    out_path = tmp_path / "out.xlsx"
    duckdb_path = csv_path.with_suffix(".duckdb")
    result = runner.invoke(
        app,
        [
            "-q",
            "-k",
            "-f",
            str(csv_path),
            "-o",
            str(out_path),
            "-r",
            "header1",
            "-c",
            "header2",
            "header3",
            "-v",
            "value",
        ],
    )
    assert result.exit_code == 0, result.output
    assert duckdb_path.exists()


def test_cli_fill_substitutes_blank_cells(tmp_path):
    """`--fill X` writes X into otherwise-empty cells."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(CSV_CONTENT)
    out_path = tmp_path / "out.xlsx"
    result = runner.invoke(
        app,
        [
            "-q",
            "--fill",
            "—",
            "-f",
            str(csv_path),
            "-o",
            str(out_path),
            "-r",
            "header1",
            "-c",
            "header2",
            "header3",
            "-v",
            "value",
        ],
    )
    assert result.exit_code == 0, result.output
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Crosstab"]
    # The (A, 2/2018) cell is empty in the source data; with --fill set, it
    # should now hold the placeholder string instead of being blank.
    assert ws["D4"].value == "—"
    wb.close()


# ── Identifier quoting & special-character headers ────────────────────────────


@pytest.mark.parametrize(
    ("raw", "quoted"),
    [
        ("simple", '"simple"'),
        ("with space", '"with space"'),
        ("Depth Interval (ft)", '"Depth Interval (ft)"'),
        ('has"quote', '"has""quote"'),
        ("café", '"café"'),
        ("123leading", '"123leading"'),
        ("select", '"select"'),
        ("group by", '"group by"'),
    ],
)
def test_quote_ident(raw, quoted):
    """Identifier quoting must double internal double-quotes and wrap in quotes."""
    assert _quote_ident(raw) == quoted


SPECIAL_HEADERS_CSV = (
    'Loc Name,Depth Interval (ft),"Param ""raw""",café,123start,select\n'
    "A,0-2,pH,arabica,x,a\n"
    "A,2-4,pH,arabica,x,b\n"
    "B,0-2,pH,robusta,x,c\n"
    "B,2-4,pH,robusta,x,d\n"
)


def test_special_character_headers_round_trip(tmp_path):
    """Headers with spaces, parentheses, embedded quotes, unicode, leading
    digits, and SQL reserved words must flow through end to end."""
    csv_path = tmp_path / "special.csv"
    csv_path.write_text(SPECIAL_HEADERS_CSV, encoding="utf-8")
    out_path = tmp_path / "special.xlsx"

    Crosstab(
        incsv=csv_path,
        outxlsx=out_path,
        row_headers=("Loc Name",),
        col_headers=("Depth Interval (ft)",),
        value_cols=('Param "raw"', "café", "123start", "select"),
    ).crosstab()

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Crosstab"]
    # The column-header label and row-header label share column A.
    assert ws["A1"].value == "Depth Interval (ft)"
    assert ws["A2"].value == "Loc Name"
    # Distinct depth-interval values are placed as column-header values
    # (each merged across the 4 value-columns).
    assert ws["B1"].value == "0-2"
    assert ws["F1"].value == "2-4"
    # Value-column labels appear in the value-label row beneath the col headers.
    value_labels = [ws.cell(row=2, column=c).value for c in range(2, 10)]
    assert value_labels == [
        'Param "raw"',
        "café",
        "123start",
        "select",
        'Param "raw"',
        "café",
        "123start",
        "select",
    ]
    # Row-header values appear in the data column.
    assert ws["A3"].value == "A"
    assert ws["A4"].value == "B"
    wb.close()


# ── Benchmark ─────────────────────────────────────────────────────────────────


def _generate_large_csv(path: Path, n_rows: int, n_cols: int) -> None:
    """Write a CSV with ``n_rows`` row keys and ``n_cols`` column keys."""
    with path.open("w", encoding="utf-8") as f:
        f.write("row_key,col_key,value\n")
        for r in range(n_rows):
            for c in range(n_cols):
                f.write(f"r{r:06d},c{c:04d},{r * n_cols + c}\n")


def test_benchmark_pivot(benchmark, tmp_path):
    """Benchmark a 1,000,000-cell crosstab end to end.

    Skipped by default via ``--benchmark-skip`` in ``pyproject.toml``.
    Run explicitly with ``uv run pytest --benchmark-only`` (or
    ``--benchmark-enable``) to time it.
    """
    n_rows, n_cols = 10_000, 100  # 1,000,000 cells total
    csv_path = tmp_path / "bench.csv"
    _generate_large_csv(csv_path, n_rows, n_cols)
    out_path = tmp_path / "bench.xlsx"

    def _run() -> None:
        Crosstab(
            incsv=csv_path,
            outxlsx=out_path,
            row_headers=("row_key",),
            col_headers=("col_key",),
            value_cols=("value",),
        ).crosstab()

    benchmark.pedantic(_run, rounds=1, iterations=1)
    assert out_path.exists()
